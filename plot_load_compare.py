#!/usr/bin/env python3
"""绘制市场披露预测统调负荷 vs 实际统调负荷（PNG）。

依赖第三方库：openpyxl + matplotlib
"""

from __future__ import annotations

import argparse
import math
from pathlib import Path

from openpyxl import load_workbook
import matplotlib.pyplot as plt
from matplotlib import font_manager


def configure_chinese_font() -> str | None:
    """配置 matplotlib 中文字体，返回命中的字体名。"""
    candidates = [
        "Microsoft YaHei",
        "SimHei",
        "Noto Sans CJK SC",
        "Noto Sans SC",
        "WenQuanYi Zen Hei",
        "PingFang SC",
        "Source Han Sans SC",
        "Arial Unicode MS",
    ]
    available = {f.name for f in font_manager.fontManager.ttflist}
    for name in candidates:
        if name in available:
            plt.rcParams["font.sans-serif"] = [name] + list(plt.rcParams.get("font.sans-serif", []))
            plt.rcParams["axes.unicode_minus"] = False
            return name
    return None


def read_series(xlsx_path: Path, sheet_name: str, key_col_idx: int, key_name: str) -> tuple[list[str], list[float]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"未找到工作表: {sheet_name}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"工作表为空: {sheet_name}")

    header = [str(v).strip() if v is not None else "" for v in rows[0]]
    time_cols = header[2:98]

    for row in rows[1:]:
        if row is None:
            continue
        key = "" if len(row) <= key_col_idx or row[key_col_idx] is None else str(row[key_col_idx]).strip()
        if key == key_name:
            vals: list[float] = []
            for i in range(2, 98):
                cell = row[i] if i < len(row) else None
                if cell is None or str(cell).strip() == "":
                    vals.append(math.nan)
                    continue
                try:
                    vals.append(float(cell))
                except (TypeError, ValueError):
                    # 兼容类似 "92,177.8" 的文本数字
                    vals.append(float(str(cell).replace(",", "").strip()))
            return time_cols, vals

    raise ValueError(f"未找到指标: {key_name}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--pred-file", default="披露数据4.2更新/信息披露查询预测信息(2026-04-03).xlsx")
    parser.add_argument("--actual-file", default="披露数据4.2更新/信息披露查询实际信息(2026-04-01).xlsx")
    parser.add_argument("--output", default="统调负荷_预测_vs_实际.png")
    args = parser.parse_args()
    selected_font = configure_chinese_font()

    times, pred_series = read_series(
        Path(args.pred_file),
        "负荷预测信息(2026-04-03)",
        key_col_idx=1,
        key_name="统调负荷(MW)",
    )

    _, actual_series = read_series(
        Path(args.actual_file),
        "机组出力情况(2026-04-01)",
        key_col_idx=0,
        key_name="统调系统实际负荷(MW)",
    )

    plt.figure(figsize=(16, 7))
    plt.plot(times, pred_series, linestyle="--", linewidth=2, label="预测统调负荷")
    plt.plot(times, actual_series, linestyle="-", linewidth=2, label="实际统调负荷")

    tick_idx = list(range(0, len(times), 8))
    if tick_idx[-1] != len(times) - 1:
        tick_idx.append(len(times) - 1)
    plt.xticks([times[i] for i in tick_idx], rotation=45)

    plt.title("市场披露统调负荷：预测 vs 实际")
    plt.xlabel("时间（15分钟）")
    plt.ylabel("负荷 (MW)")
    plt.grid(alpha=0.25)
    plt.legend()
    plt.tight_layout()

    output = Path(args.output)
    plt.savefig(output, dpi=180)
    if selected_font:
        print(f"已生成PNG图表: {output}（中文字体: {selected_font}）")
    else:
        print(
            "已生成PNG图表: "
            f"{output}（未检测到常见中文字体，请安装如 Noto Sans CJK SC / Microsoft YaHei / SimHei）"
        )


if __name__ == "__main__":
    main()
